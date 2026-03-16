"""
ropa_parser.py
==============
Parses both ROPA Excel formats used by the organisation:

FORMAT A – Vertical (RoPA_Test.xlsx style)
  One process per sheet.  Col B = field label, Col C = value.
  Detected when cell B2 or B3 contains "Data Fiduciary".

FORMAT B – Horizontal (RoPA_Template.xlsx style)
  Sheet named "RoPA".  Row 12 = column headers, row 13+ = one process per row.
  Sections 1-7 span 53 columns.
"""

import io
import openpyxl
from typing import List, Dict, Any


# ── FORMAT A helpers ───────────────────────────────────────────────────────────

_FORMAT_A_MAP = {
    # label keyword (lowercase)  →  canonical key
    "business function":                     "function_name",
    "purpose of processing":                 "purpose",
    "categories of data principals":         "data_subjects",
    "categories of personal data":           "personal_data_categories",
    "categories of internal recipients":     "internal_recipients",
    "categories of external recipients":     "external_recipients",
    "list of third-party vendors":           "third_party_vendors",
    "jurisdiction of processing":            "jurisdiction_processing",
    "jurisdiction of data principals":       "jurisdiction_data_principals",
    "safeguards for exceptional transfers":  "transfer_safeguards",
    "retention period":                      "retention_period",
    "general description of technical":      "security_measures",
    "lawful basis":                          "lawful_basis",
    "legitimate uses":                       "legitimate_uses",
    "rights available":                      "data_subject_rights",
    "automated decision":                    "automated_decision_making",
    "source of the personal data":           "data_sources",
    "link to record of consent":             "consent_record_link",
    "location of personal data":             "data_location",
    "data protection impact assessment required": "dpia_required",
    "data protection impact assessment progress": "dpia_progress",
    "has a personal data breach occurred":   "breach_occurred",
    "company name":                          "company_name",
    "name and contact":                      "_skip",
}


def _parse_format_a(ws) -> Dict[str, Any]:
    """Extract key-value pairs from a vertical-layout sheet."""
    proc: Dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        label = str(row[1]).strip() if row[1] else ""
        value = str(row[2]).strip() if row[2] else ""
        if not label or label == "None":
            continue
        label_lower = label.lower()
        matched_key = None
        for keyword, key in _FORMAT_A_MAP.items():
            if keyword in label_lower:
                matched_key = key
                break
        if matched_key == "_skip":
            continue
        if matched_key:
            proc[matched_key] = value
        # capture company info from header rows
        if "company name" in label_lower and value:
            proc["company_name"] = value
    return proc


# ── FORMAT B helpers ───────────────────────────────────────────────────────────

# Mapping from column index (0-based in the RoPA sheet) to canonical key
_FORMAT_B_COL_KEYS = {
    0:  "country",
    1:  "function_name",
    2:  "process_name",
    3:  "process_owner",
    4:  "purpose",
    5:  "data_subjects",
    6:  "purpose_detail",
    7:  "data_controller_entity",
    8:  "entity_role",
    9:  "personal_data_categories",
    10: "pii_direct_collection",
    11: "public_source_collection",
    12: "third_party_data_received",
    13: "application_data_collection",
    14: "internal_team_data",
    15: "privacy_notice_provided",
    16: "consent_obtained",
    17: "lawful_basis",
    18: "manual_processing",
    19: "application_software_used",
    20: "storage_location",
    21: "hosting_type",
    22: "vendor_access_list",
    23: "vendor_access_locations",
    24: "data_principal_rights_mechanism",
    25: "role_based_access",
    26: "user_access_review",
    27: "encryption",
    28: "authentication_mechanism",
    29: "password_policy",
    30: "mobile_device_separation",
    31: "two_factor_auth",
    32: "security_measures",
    33: "storage_laptop_sharepoint",
    34: "hard_copy_storage",
    35: "cloud_storage",
    36: "retention_requirement",
    37: "retention_period",
    38: "retention_policy_exists",
    39: "data_disposal_method",
    40: "internal_transfer",
    41: "internal_transfer_recipient",
    42: "transfer_purpose",
    43: "transfer_mode",
    44: "cross_border_transfer",
    45: "cross_border_mechanism",
    46: "third_party_disclosure",
    47: "vendor_location",
    48: "data_sent_across_border",
    49: "sharing_purpose",
    50: "shared_pii_categories",
    51: "sharing_mode",
    52: "contract_dpa_exists",
}


def _parse_format_b(ws) -> List[Dict[str, Any]]:
    """Extract one dict per data row from a horizontal-layout RoPA sheet."""
    processes = []
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Row 12 (0-indexed 11) contains the headers — detect by "Country" in col 0
        if row[0] is not None and "Country" in str(row[0]):
            header_row_idx = i
            continue
        if header_row_idx is None:
            continue
        # Skip rows that are entirely empty
        if all(c is None for c in row):
            continue
        # Skip section-label rows (merged header rows)
        if isinstance(row[0], str) and "Section" in str(row[0]):
            continue
        if isinstance(row[0], (int, float)) and row[1] is None:
            continue

        proc: Dict[str, Any] = {}
        for col_idx, key in _FORMAT_B_COL_KEYS.items():
            if col_idx < len(row):
                val = row[col_idx]
                proc[key] = str(val).strip() if val is not None else ""
        # Only keep rows that have at least a function or process name
        if proc.get("function_name") or proc.get("process_name"):
            processes.append(proc)

    return processes


# ── Public API ─────────────────────────────────────────────────────────────────

def parse_ropa_excel(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parse a ROPA Excel file.  Returns a list of process dicts,
    one per processing activity found.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    processes: List[Dict[str, Any]] = []

    # ── Detect Format B first (has a sheet called "RoPA") ──────────────────
    if "RoPA" in wb.sheetnames:
        ws  = wb["RoPA"]
        raw = _parse_format_b(ws)
        if not raw:
            # Template detected but no data rows filled in
            return [{
                "_format":  "B_EMPTY",
                "id":       "TEMPLATE",
                "process_name": "ROPA Template — No data rows found",
                "notes": (
                    "This file is the blank ROPA_Template. "
                    "Fill in the RoPA sheet (row 13 onwards) then re-upload."
                ),
            }]

        # Enrich with Glossary data if present
        glossary: Dict[str, List[str]] = {}
        if "Glossary" in wb.sheetnames:
            gws = wb["Glossary"]
            cols: List[List[str]] = [[], [], [], []]
            for row in gws.iter_rows(min_row=4, values_only=True):
                for j in range(4):
                    if j < len(row) and row[j]:
                        cols[j].append(str(row[j]))
            glossary = {
                "personal_data_elements":      cols[0],
                "data_principal_categories":   cols[1],
                "lawful_basis_options":        cols[2],
                "entity_roles":                cols[3],
            }

        for proc in raw:
            proc["_format"]  = "B"
            proc["glossary"] = glossary
            proc["id"]       = f"P{len(processes)+1:03d}"
            processes.append(proc)
        return processes

    # ── Format A: iterate all sheets (also fallback if Format B was empty) ─
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Detect vertical format: look for "Data Fiduciary" in first 5 rows
        is_format_a = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 5:
                break
            for cell in row:
                if cell and "Data Fiduciary" in str(cell):
                    is_format_a = True
                    break

        if is_format_a:
            proc = _parse_format_a(ws)
            if proc:
                proc["_format"] = "A"
                proc["id"]      = f"P{len(processes)+1:03d}"
                proc["sheet"]   = sheet_name
                processes.append(proc)

    return processes


def processes_to_text(processes: List[Dict[str, Any]]) -> str:
    """Convert list of process dicts to a readable text block for AI prompts."""
    lines = []
    for p in processes:
        lines.append(f"\n{'='*60}")
        lines.append(f"PROCESS ID   : {p.get('id','?')}")
        for k, v in p.items():
            if k.startswith("_") or k == "glossary" or not v or v in ("None",""):
                continue
            lines.append(f"{k.replace('_',' ').title():35s}: {v}")
    return "\n".join(lines)
